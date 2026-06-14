############################################################
#
#     PROYECTO: WEB SCRAPING DE MERCADO PUBLICO CHILE
#     POR: ERICK INFANTE
#     FECHA: 30-05-2024
#     LICITACIONES
#
############################################################

from datetime import date

import requests
import pandas as pd
import zipfile
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import unicodedata
from tkinter.simpledialog import askstring
from tkinter import Tk
import re
from pathlib import Path
import os

from logger_config import setup_logger
import traceback

logger = setup_logger()

# URL del archivo a descargar
url: str = "https://www.mercadopublico.cl/Portal/att.ashx?id=5"

BASE_DIR: Path = Path(__file__).parent
DOWNLOAD_DIR: Path = BASE_DIR / "downloads"
DOWNLOAD_DIR.mkdir(exist_ok=True)

today: pd.Timestamp = pd.Timestamp.today()


# Función para descargar y descomprimir el archivo
def descargar_y_descomprimir(url: str, destfile: str, exdir: Path) -> None:
    """
    Descarga un archivo ZIP y lo descomprime
    """
    logger.info("Descargando archivo...")
    try:
        # Hacer la solicitud HTTP para descargar el archivo
        response = requests.get(url)

        # Guardar el archivo descargado en el sistema
        with open(destfile, "wb") as f:
            f.write(response.content)

        # Descomprimir el archivo ZIP
        with zipfile.ZipFile(destfile, "r") as zip_ref:
            zip_ref.extractall(exdir)

    except Exception as e:
        logger.error(
            f"Error al descargar o descomprimir el archivo: {e} | traceback: {traceback.format_exc()}"
        )
        raise


# Función para eliminar acentos
def remove_accents(text: str) -> str:
    """
    Elimina acentos del texto
    """
    #  logger.info(f"Eliminando acentos de texto...")
    try:
        if isinstance(text, str):
            text = unicodedata.normalize("NFKD", text)
            text = "".join([c for c in text if not unicodedata.combining(c)])
        return text

    except Exception as e:
        logger.error(
            f"Error al eliminar acentos: {e} | traceback: {traceback.format_exc()}"
        )
        raise


def manejo_datos() -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Maneja los datos de licitaciones: descarga, descomprime, lee el Excel, limpia y filtra los datos.
    """
    logger.info("Manejando datos de licitaciones...")
    try:
        # Llamar a la función
        descargar_y_descomprimir(
            url, DOWNLOAD_DIR / "licitaciones.zip", DOWNLOAD_DIR / "licitaciones"
        )

        # Leer el archivo Excel descomprimido, omitiendo las primeras 7 filas
        licitaciones: pd.DataFrame = pd.read_excel(
            DOWNLOAD_DIR / "licitaciones" / "Licitacion_Publicada.xlsx", skiprows=7
        )

        # Eliminar columnas innecesarias (columnas en posición 2, 7, 8, 11-13 en indexación 0-based)
        licitaciones = licitaciones.drop(licitaciones.columns[[0, 3, 8, 12]], axis=1)

        # Cambiar nombre de columnas (Eliminar acentos)
        licitaciones.columns = [remove_accents(col) for col in licitaciones.columns]

        # Cambiar formato de columnas
        licitaciones["Fecha Cierre"] = pd.to_datetime(
            licitaciones["Fecha Cierre"]
        )  # , origin=pd.Timestamp('1899-12-30'), unit='D')
        licitaciones["Region Compradora"] = licitaciones["Region Compradora"].astype(
            "category"
        )
        licitaciones["Organismo"] = licitaciones["Organismo"].astype("category")

        # Filtrar solo registros de licitaciones abiertas (fecha de cierre > Hoy)
        licitaciones = licitaciones[licitaciones["Fecha Cierre"] > today]

        # Filtrar por región
        licitaciones_maule: pd.DataFrame = licitaciones[
            licitaciones["Region Compradora"].str.contains("Maule")
        ]

        return licitaciones, licitaciones_maule
    except Exception as e:
        logger.error(
            f"Error al manejar los datos: {e} | traceback: {traceback.format_exc()}"
        )
        raise


def guardar_excel(licitaciones, licitaciones_maule):
    logger.info("Guardando datos en Excel...")
    try:
        # Crear un objeto workbook
        wb = Workbook()

        # Crear la primera hoja (Maule) y agregar datos
        if "Maule" in wb.sheetnames:
            ws = wb["Maule"]
        else:
            ws = wb.active
            ws.title = "Maule"

        # Escribir datos de la región del Maule
        for r_idx, row in enumerate(
            dataframe_to_rows(licitaciones_maule, index=False, header=True), 1
        ):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        # Obtener palabras para filtrar mediante diálogo
        root = Tk()
        root.withdraw()
        respuesta = askstring(
            "Filtrar por palabras",
            "Introduce las palabras para filtrar (separadas por comas):",
            initialvalue="parka, polera, polerón, micropolar, guante, nitrilo, anti corte, casco, botín, bota, calzado, ropa, seguridad, capas, pantalón, gorro, arnés, mascarilla, filtro, cabritilla, acondicionado",
        )
        root.destroy()

        if respuesta:
            palabras = [palabra.strip() for palabra in respuesta.split(",")]

            # Bucle para filtrar por cada palabra y agregar una hoja al workbook
            for palabra in palabras:
                # Filtrar los registros donde la palabra sea encontrada
                filtradas = licitaciones[
                    licitaciones.apply(
                        lambda row: any(
                            re.search(palabra, str(cell), re.IGNORECASE) for cell in row
                        ),
                        axis=1,
                    )
                ]

                # Agregar una hoja al workbook con las licitaciones filtradas
                if len(filtradas) > 0:
                    ws = wb.create_sheet(palabra)
                    for r_idx, row in enumerate(
                        dataframe_to_rows(filtradas, index=False, header=True), 1
                    ):
                        for c_idx, value in enumerate(row, 1):
                            ws.cell(row=r_idx, column=c_idx, value=value)

            # Eliminar el archivo ZIP descargado
            zip_path = DOWNLOAD_DIR / "licitaciones.zip"
            if zip_path.exists():
                zip_path.unlink()

            # Guardar el workbook en un archivo Excel
            v = 1
            while Path(
                f"C://Users//{os.getlogin()}//Downloads//licitaciones - {date.today().strftime('%d-%m-%Y')} - {v}.xlsx"
            ).exists():
                v += 1
            archivo_salida = Path(
                f"C://Users//{os.getlogin()}//Downloads//licitaciones - {date.today().strftime('%d-%m-%Y')} - {v}.xlsx"
            )

            wb.save(archivo_salida)
            logger.info(f"Archivo guardado: {archivo_salida}")
        else:
            logger.info("Operación cancelada por el usuario")
    except Exception as e:
        logger.error(
            f"Error al guardar el archivo Excel: {e} | traceback: {traceback.format_exc()}"
        )
        raise


def main():
    """Función principal para ejecutar el proceso completo"""
    logger.info("Iniciando proceso de manejo de licitaciones...")
    try:
        licitaciones, licitaciones_maule = manejo_datos()
        guardar_excel(licitaciones, licitaciones_maule)
    except Exception as e:
        logger.error(
            f"Error en el proceso principal: {e} | traceback: {traceback.format_exc()}"
        )
        logger.info(
            "Ocurrió un error durante el proceso. Revisa el log para más detalles."
        )


if __name__ == "__main__":
    main()
